from flask import Flask, request, send_file, render_template
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import tempfile
import os

app = Flask(__name__)

@app.route('/', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        if 'file' not in request.files:
            return render_template('index.html', error="No file selected")
            
        file = request.files['file']
        if file.filename == '':
            return render_template('index.html', error="No file selected")
            
        if file and file.filename.endswith('.xlsx'):
            # Save uploaded file
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_input:
                file.save(tmp_input.name)
                
            # Process the file
            try:
                output_path = process_file(tmp_input.name)
                return send_file(output_path, as_attachment=True, download_name='ajustado_' + file.filename)
            finally:
                os.remove(tmp_input.name)
                os.remove(output_path)
                
    return render_template('index.html')

def process_file(input_path):
    wb = load_workbook(input_path)
    ws = wb.active
    row_number = 11

    for _ in ws.iter_rows(min_row=row_number, max_row=ws.max_row):
        cell = ws[f'G{row_number}']
        if cell.value is not None:
            decimal_hours = float(str(cell.value).replace(',', '.'))
            hours = int(decimal_hours)
            minutes = int(round((decimal_hours - hours) * 60))
            ws[f'H{row_number}'] = f"{hours:02}:{minutes:02}"
            ws[f'H{row_number}'].alignment = Alignment(horizontal='right')
        row_number += 1

    # Save to temporary output
    tmp_output = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(tmp_output.name)
    return tmp_output.name

if __name__ == '__main__':
    app.run(debug=False)
